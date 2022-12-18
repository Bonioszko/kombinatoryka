import random
import copy
import itertools
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
import numpy as np
import time
import openpyxl


def distance_graph_generate(distance_min, distance_max, n):
    # generowanie macierzy dystansu pomiedzy poszczególnymi punktami
    distances = []
    for i in range(n):
        distances.append([0]*n)
    for i in range(n-1):
        for j in range(i+1, n):
            distances[j][i] = distances[i][j] = random.randint(
                distance_min, distance_max)
    return distances

# wzor na obliczanie czasu przejazdu to s/v


def distance_from_base_generate(distance_min, distance_max, n):
    distances = []
    for i in range(n):
        distances.append(random.randint(distance_min, distance_max))
    return distances


def times_from_base_generate(distances, speed):
    times = copy.deepcopy(distances)
    n = len(distances)
    for i in range(n):
        times[i] = distances[i] / speed*60
    return times


def time_graph_generate(speed, distances):
    # generacja macierzy czasu potrzebnego na przejechanmie czas w godzinach,
    times = copy.deepcopy(distances)
    for i in range(len(distances)-1):
        for j in range(i+1, len(distances)):
            times[i][j] = times[j][i] = round(distances[i][j]/speed * 60, 0)

    return times


# czas będziemy podawać w minutach od rozpączecia dnia, pierwszy czas to poczatek dostęności, drugi to koniec, czyli np zakres(300,660) odpowiada dostępności w godzinach 5:00 do 11:00
# zeby algorymt mial szanse sie wykonac, to zakladamy, ze ktos jest w domu przez 500 min
def time_stamps_generate(n, min_time, max_time, time_avaliable):
    time_stamps = []
    for i in range(n):
        time_stamps.append([0, 0])
    for i in range(n):
        a = random.randint(min_time, max_time)
        time_stamps[i][0] = a
        time_stamps[i][1] = a+time_avaliable
    return time_stamps
# do generacji permutacji uzyjemy wbudowanej biblioteki pythopa
# tutaj brute force, który wygeneruje wszystkie możliwe rozwiązania i sprawdzi przy którym pokonamy najmniejszy dystanms (parametry to wszystkie grpahy które wygenerowalismy + start i koniec pracy )


# ten nie dziala
def brute_force_min_dist(time_stamps, distances, times, start, end):
    n = len(time_stamps)
    best_road = [0]*n
    min_dist = 100000000
    # generuje tutaj wszystkie mozliwe permutacje listy o dlugosci n
    iterations = list(itertools.permutations(list(range(n))))

    for i in iterations:
        d = 0
        dist = 0
        time = start
        for j in range(n-1):
            dist += distances[i[j]][i[j+1]]
            time += times[i[j]][i[j+1]]
        dist += distances[i[-1]][i[0]]
        time += times[i[-1]][i[0]]

        if dist < min_dist:
            min_dist = dist
            best_road = i
    best_road = list(best_road)
    for i in range(n):
        best_road[i] += 1

    return min_dist, best_road


def brute_force_min_time(base_times, base_distances, time_stamps, distances, times, start, end):
    n = len(time_stamps)
    best_road = [0]*n
    min_dist = 100000000
    work_time = end-start
    min_time = 1000000
    # generuje tutaj wszystkie mozliwe permutacje listy o dlugosci n
    iterations = list(itertools.permutations(list(range(n))))
    print(base_times, end="\n")
    print(base_distances, end="\n")

    for i in iterations:
        d = end
        dist = 0
        time = start
        # div_start = time_stamps[i[0]][0] - time - base_times[i[0]]
        # if div_start > 0:
        #     time += div_start
        # if time_stamps[i[0]][1] > start + base_times[i[0]]:

        #     dist += base_distances[i[0]]
        #     time += base_times[i[0]]
        for j in range(n-1):
            div = time_stamps[i[j+1]][0] - time - times[i[j]][i[j+1]]
            if div > 0:
                time += div
            if time_stamps[i[j+1]][1] < times[i[j]][i[j+1]]+time:
                continue
            dist += distances[i[j]][i[j+1]]
            time += times[i[j]][i[j+1]]
        dist += distances[i[-1]][i[0]]
        time += times[i[-1]][i[0]]
        # dist += base_distances[i[-1]]
        # time += base_times[i[-1]]
        if time > end:
            continue

        elif time-start < min_time:
            min_time = time-start
            best_road = i

    return min_time, best_road


def brute_force_min_dist2(time_stamps, distances, times, start, end):
    n = len(time_stamps)
    best_road = [0]*n
    min_dist = 100000000
    iterations = list(itertools.permutations(list(range(n))))
    for i in iterations:
        d = 0
        dist = 0
        time = start
        for j in range(n-1):
            div = time_stamps[i[j+1]][0] - time - times[i[j]][i[j+1]]
            if div > 0:
                time += div
            if time_stamps[i[j+1]][1] < times[i[j]][i[j+1]]+time:
                continue
            d += 1
            dist += distances[i[j]][i[j+1]]
            time += times[i[j]][i[j+1]]
        dist += distances[i[-1]][i[0]]
        time += times[i[-1]][i[0]]
        if time > end:
            continue
        elif dist < min_dist and d == n-1:
            min_dist = dist
            best_road = i
    best_road = list(best_road)
    if len(best_road) == n:
        for i in range(n):
            best_road[i] += 1

    return min_dist, best_road


def return_min_value_column(array, visited, time, times, current):
    index_min = 10000
    min = 100000
    div_end = 0
    time_help = time
    for i in range(len(array)):
        time = time_help
        div = array[i][0]-time - times[current][i]
        if div > 0:
            time += div
        if array[i][0] < min and visited[i] == 0 and array[i][1]-time-times[current][i] > 0:
            index_min = i
            min = array[i][0]
            div_end = div
    return index_min, div_end
# po wszystkich tych ktorzy sa najwczesniej dostępni, wyniki nawet maja sens


def shortest_time(time_stamps, distances, times, start, end):
    n = len(distances)
    visited = [0]*n
    best_path = [0]*n
    start_index = 10000
    distance = 1000
    for k in range(n):
        path = []
        dist = 0
        time = start
        for j in range(n):
            visited[j] = 0
        visited[k] = 1
        start_index = k
        current = k
        for i in range(n-1):
            visited[current] = 1
            next, div = return_min_value_column(
                time_stamps, visited, time, times, current)
            if next != 10000:
                dist += distances[current][next]
                time += div
                time += times[current][next]
                path.append(current)
                current = next
        if time + times[current][-1] < end:
            path.append(current)
            dist += distances[current][start_index]
        if dist < distance:
            distance = dist
            best_path = path
    if len(best_path) == n:
        for i in range(n):
            best_path[i] += 1
    return distance, best_path


def return_min_time_to_live(time_stamps, visited, time, current, times):
    index_min = 10000
    min = 1000000
    div_end = 0
    time_help = time
    for i in range(len(time_stamps)):
        time = time_help
        div = time_stamps[i][0]-time - times[current][i]
        if div > 0:
            time += div
        if time_stamps[i][1]-time < min and visited[i] == 0 and time_stamps[i][1]-time - times[current][i] > 0:
            min = time_stamps[i][1]-time
            index_min = i
            div_end = div
    return index_min, div_end


def shortest_time_to_live(time_stamps, distances, times, start, end):
    n = len(distances)
    visited = [0]*n
    best_path = [0]*n
    start_index = 10000
    distance = 1000
    for k in range(n):
        path = []
        dist = 0
        time = start
        for j in range(n):
            visited[j] = 0
        visited[k] = 1
        start_index = k
        current = k
        for i in range(n-1):
            visited[current] = 1
            next, div = return_min_time_to_live(
                time_stamps, visited, time, current, times)
            if next != 10000:
                dist += distances[current][next]
                time += div
                time += times[current][next]
                path.append(current)
                current = next
        if time + times[current][-1] < end:
            path.append(current)
            dist += distances[current][start_index]

        if dist < distance and len(path) == n:
            distance = dist
            best_path = path
    if len(best_path) == n:
        for i in range(n):
            best_path[i] += 1
    return distance, best_path
# chyba dziala

# to dosc wazne na dole

# def test_na_ilosc(n, speed):
#     distances = distance_graph_generate(10, 70, n)
#     times = time_graph_generate(speed, distances)
#     time_stamps = time_stamps_generate(n, 200, 900, 200)
#     print(times)
#     print(distances)
#     print(time_stamps)
#     distancesss, best_roadsss = brute_force_min_dist2(
#         time_stamps, distances, times, 100, 1440)
#     print(distancesss)
#     print(best_roadsss)
#     distances1, best_road1 = brute_force_min_dist(
#         time_stamps, distances, times, 100, 1440)
#     print(distances1)
#     print(best_road1)
#     distances_sh_time, path_sh_time = shortest_time(
#         time_stamps, distances, times, 100, 1440)
#     distances_leave, path_leave = shortest_time_to_live(
#         time_stamps, distances, times, 100, 1440)
#     print(distances_sh_time)
#     print(path_sh_time)
#     print(distances_leave)
#     print(path_leave)
#     # Create a graph from the distance matrix
#     my_array = np.matrix(distances)
#     # Create a graph from the distance matrix
#     G = nx.from_numpy_matrix(my_array)
#     # Use the spring layout to position the nodes
#     pos = nx.spring_layout(G)
#     # Add labels to the nodes
#     labels = {
#         i: f" nr:{i} {time_stamps[i][0]}:{time_stamps[i][1]}" for i in G.nodes()}
#     nx.draw_networkx_labels(G, pos, labels, font_size=20)
#     indices = np.triu_indices_from(my_array)
#     # Create pairs of indices
#     pairs = list(zip(indices[0], indices[1]))
#     for i in range(len(pairs)):
#         t = list(pairs[i])
#         pairs[i] = t
#     # Use the pairs of indices to index the distances list
#     edges1 = []
#     for i, j in pairs:
#         if i != j:
#             edges1.append(my_array[i, j])
#     edge_labels = dict(zip(G.edges(), edges1))
#     nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=20)
#     # Draw the graph
#     nx.draw(G, pos)
#     # Show the plot
#     plt.show()


# test_na_ilosc(5, 50)


def vizualization(n, speed):
    distance_list1 = []
    distance_list2 = []
    distance_list3 = []
    distance_list4 = []
    edges = []
    for i in range(5, 11):
        distances = distance_graph_generate(10, 70, i)
        times = time_graph_generate(speed, distances)
        time_stamps = time_stamps_generate(i, 200, 900, 200+(i*100))
        start = time.time()
        distancesss, best_roadsss = brute_force_min_dist2(
            time_stamps, distances, times, 100, 1440 + (i*100))
        end = time.time()
        distance_list1.append(end-start)
        # distances1, best_road1 = brute_force_min_dist(
        #     time_stamps, distances, times, 100, 1440 + (i*100))
        start = time.time()
        distances_sh_time, path_sh_time = shortest_time(
            time_stamps, distances, times, 100, 1440 + (i*100))
        end = time.time()
        distance_list3.append(end-start)
        start = time.time()
        distances_leave, path_leave = shortest_time_to_live(
            time_stamps, distances, times, 100, 1440 + (i*100))
        end = time.time()
        distance_list4.append(end-start)
        # # distance_list2.append(distances1)
        # distance_list3.append(distances_sh_time)
        # distance_list4.append(distances_leave)
        edges.append(i)
    fig = plt.figure()
    fig.suptitle('`distances')
    # Add a subplot
    ax = fig.add_subplot(111)
    # Plot the first line and set the label
    ax.plot(edges, distance_list1, 'r--', label='brut_force', marker='.')
    print(distance_list1)
    # ax.plot(edges, distance_list2, 'g--', label='second')
    ax.plot(edges, distance_list3, 'g--',  label='first_greedy', marker='.')
    ax.plot(edges, distance_list4, 'b--', label='second_greedy', marker='.')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, item in enumerate(distance_list1):
        sheet.cell(row=i+1, column=1).value = item
    for i, item in enumerate(distance_list3):
        sheet.cell(row=i+1, column=2).value = item
    for i, item in enumerate(distance_list4):
        sheet.cell(row=i+1, column=3).value = item
    workbook.save('my_list.xlsx')
    ax.legend(loc="upper left")
    ax.set_xlabel('number of vertexes')
    ax.set_ylabel('time to execute(s)')
    plt.show()


vizualization(5, 50)

# tutaj dopisze najblizsze czas, od aktualnego (moze zadziala)
# n = 8
# speed = 100
# distances = distance_graph_generate(2, 10, n)
# times = time_graph_generate(20, distances)
# time_stamps = time_stamps_generate(n, 200, 900, 500)
# time_stamps = []
# for i in range(n):
#     time_stamps.append([100, 1200])
# print(time_stamps)
# base_distances = distance_from_base_generate(20, 50, n)
# base_times = times_from_base_generate(base_distances, speed)

# print(list(itertools.permutations(list(range(n)))))
# print(distances)
# print(times)
# print(time_stamps)
# print(distances)
# distancesss, best_roadsss = brute_force_min_dist2(
#     time_stamps, distances, times, 100, 1440)
# distances_sh_time, path_sh_time = shortest_time(
#     base_times, base_distances, time_stamps, distances, times, 100, 1440)
# distances_leave, path_leave = shortest_time_to_live(
#     base_times, base_distances, time_stamps, distances, times, 100, 1440)
# print(distancesss)
# print(best_roadsss)

# print(distances_sh_time)
# print(path_sh_time)
# print(distances_leave)
# print(path_leave)
# dwa brute forcy zrobione teraz dopisac cos jeszcze jakies heurystyki czy cos
